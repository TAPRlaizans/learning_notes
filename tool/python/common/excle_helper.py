import os
import xlrd
import xlwt
import openpyxl
from xlutils.copy import copy


from common.file_helper import FileHelper as fh

class ExcleHelper(object):
    @staticmethod
    def save_excle(data_list, sheet_name="data", path_output_folder='./output/'):
        length = len(data_list)
        path_output_file = os.path.join(path_output_folder, f'{sheet_name}.xlsx')
        fh.check_folder_exist(path_output_folder, True)
        wb = openpyxl.Workbook()
        sheet = wb.active 
        sheet.title = sheet_name
        for i in range(0, length):
            for j in range(0, len(data_list[i])):
                sheet.cell(row=i+1, column=j+1, value=str(data_list[i][j]))
        wb.save(f'{path_output_file}')

    @staticmethod
    def write_excel_xls(path, sheet_name, value, list_header=None):
        if list_header != None:
            value.insert(0, list_header)
        index = len(value)  
        workbook = xlwt.Workbook() 
        sheet = workbook.add_sheet(sheet_name)  
        for i in range(0, index):
            for j in range(0, len(value[i])):
                sheet.write(i, j, value[i][j])  
        workbook.save(path)  
        print(f"{path}表格写入数据成功！")
